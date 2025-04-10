# Created by Benjamin Richards 04/04/2025
# Titan Fittings Web Scraper - Scrapes product titles and image URLs to Excel
# MAKE SURE TO USE GLOBAL PYTHON INTERPRETER SINCE THAT IS WHERE ALL OF THE PACKAGES ARE INSTALLED CORRECTLY

import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
from urllib.parse import urljoin, unquote, urlparse, parse_qs

def scrape_product_info(url):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Referer': 'https://www.titanfittings.com/'
        }

        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Product title extraction
        title_selectors = [
            'h1.product-name',
            'h1.product-title',
            'h1.product-detail__title',
            'h1.productView-title',
            'h1[itemprop="name"]',
            'h1.title'
        ]

        product_title = None
        for selector in title_selectors:
            element = soup.select_one(selector)
            if element:
                product_title = element.get_text(strip=True)
                break

        if not product_title:
            h1 = soup.find('h1')
            if h1:
                product_title = h1.get_text(strip=True)

        if not product_title:
            raise ValueError("Could not find product title on page")

          # Next.js Image Handling
        image_url = None
        nextjs_img = soup.select_one('img[src*="/_next/image"]')

        # Method 1: Extract from src attribute
        if nextjs_img:
            src = nextjs_img.get('src', '')
            if 'url=' in src:
                parsed = urlparse(src)
                params = parse_qs(parsed.query)
                if 'url' in params:
                    image_url = unquote(params['url'][0])
                    if image_url.startswith('/'):
                        image_url = urljoin('https://www.titanfittings.com', image_url)

        # Method 2: Fallback to srcset
        if not image_url and nextjs_img and nextjs_img.get('srcset'):
            srcset = nextjs_img['srcset'].split(',')[0].strip()
            if 'url=' in srcset:
                srcset_url = srcset.split(' ')[0]
                parsed = urlparse(srcset_url)
                params = parse_qs(parsed.query)
                if 'url' in params:
                    image_url = unquote(params['url'][0])
                    if image_url.startswith('/'):
                        image_url = urljoin('https://www.titanfittings.com', image_url)

        # Method 3: Fallback to data attributes
        if not image_url:
            img = soup.select_one('img[data-src*="api.titanfittings.com"]')
            if img and img.get('data-src'):
                image_url = img['data-src']
                if image_url.startswith('/'):
                    image_url = urljoin('https://www.titanfittings.com', image_url)

        return {
            'title': product_title,
            'image_url': image_url if image_url else 'N/A',
            'product_url': url
        }

    except requests.exceptions.RequestException as e:
        print(f"Error during request to {url}: {str(e)}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while processing {url}: {str(e)}")
        return None

def save_to_excel(data, file_path):
    try:
        # Try to load existing workbook
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        except FileNotFoundError:
            # If file doesn't exist, create new workbook
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Product Title", "Image URL", "Product URL", "Date Added"])
        except Exception as e:
            print(f"Error loading/creating Excel file {file_path}: {e}")
            return

        # Write data
        sheet.append([
            data.get('title', 'N/A'),
            data.get('image_url', 'N/A'),
            data.get('product_url', 'N/A'),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

        # Save workbook
        wb.save(file_path)
        print(f"Data successfully saved to {file_path}")

    except Exception as e:
        print(f"Error saving to Excel: {e}")

if __name__ == "__main__":
    # Define paths
    links_excel_path = "BOP.xlsx"  # Input file with links
    output_excel_path = "BOP_output_urls.xlsx"  # Output file for URLs

    try:
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(links_excel_path)
        sheet = workbook.active

        # Iterate through the rows in column J (10th column)
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=10).value

            if cell_value and isinstance(cell_value, str) and cell_value.startswith("https://www.titanfittings.com/"):
                product_url = cell_value
                print(f"\nProcessing URL: {product_url}")

                # Scrape product info
                product_info = scrape_product_info(product_url)

                if product_info:
                    print(f"Scraped product info: {product_info['title']}, Image URL: {product_info['image_url']}")

                    # Save to Excel with the image URL
                    save_to_excel(product_info, output_excel_path)
                else:
                    print(f"Failed to scrape product information for {product_url}")
            elif cell_value:
                print(f"Skipping non-Titan Fittings link or empty cell in row {row}, column J: {cell_value}")
            else:
                print(f"Skipping empty cell in row {row}, column J.")

        print("\nScraping completed successfully! Image URLs saved to {output_excel_path}")

    except FileNotFoundError:
        print(f"Error: The file '{links_excel_path}' was not found in the current directory.")
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")